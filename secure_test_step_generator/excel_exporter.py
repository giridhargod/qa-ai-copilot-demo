"""Excel Export stage: GenerationResult -> .xlsx, written exactly to
the folder the user specified.
"""
import os

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font

from .models import GenerationResult

COLUMNS = [
    "Step No",
    "Action",
    "Expected Result",
    "Confidence",
    "Confidence Reason",
    "Warnings",
    "Status",
]
COLUMN_WIDTHS = [10, 45, 45, 12, 35, 35, 12]


def _excel_safe(value):
    """Strip characters the .xlsx XML format itself forbids (e.g. NUL,
    other control chars). This is real, not theoretical: garbled OCR
    output from a low-quality screenshot can produce these, and
    openpyxl raises IllegalCharacterError and aborts the *entire*
    export over a single bad cell otherwise.
    """
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def export_to_excel(result: GenerationResult, output_dir: str, filename: str | None = None) -> str:
    os.makedirs(output_dir, exist_ok=True)

    if filename is None:
        stem = os.path.splitext(os.path.basename(result.source_document))[0]
        filename = f"{stem}_test_steps.xlsx" if stem else "test_steps.xlsx"

    output_path = os.path.join(output_dir, filename)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Test Steps"

    sheet.append(COLUMNS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for step in result.steps:
        sheet.append(
            [
                step.step_no,
                _excel_safe(step.action),
                _excel_safe(step.expected_result),
                f"{round(step.confidence * 100)}%",
                _excel_safe(step.confidence_reason),
                _excel_safe("\n".join(step.warnings)),
                step.status,
            ]
        )

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for index, width in enumerate(COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[sheet.cell(row=1, column=index).column_letter].width = width

    workbook.save(output_path)
    return output_path
