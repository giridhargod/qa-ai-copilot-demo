import os
import tempfile

from openpyxl import load_workbook

from ..excel_exporter import export_to_excel
from ..models import GenerationResult, TestStep


def _sample_result():
    return GenerationResult(
        source_document="C:/docs/sample.docx",
        steps=[
            TestStep(
                step_no=1,
                action="Click Submit",
                expected_result="Confirmation page appears",
                confidence=0.84,
                confidence_reason="OCR quality high; cursor detected",
                warnings=["One ambiguous UI element"],
            )
        ],
    )


def test_writes_xlsx_to_exact_output_folder_with_expected_columns():
    with tempfile.TemporaryDirectory() as tmp_dir:
        output_path = export_to_excel(_sample_result(), tmp_dir)

        assert os.path.dirname(output_path) == tmp_dir
        assert os.path.exists(output_path)

        wb = load_workbook(output_path)
        sheet = wb.active

        header = [cell.value for cell in sheet[1]]
        assert header == [
            "Step No", "Action", "Expected Result",
            "Confidence", "Confidence Reason", "Warnings", "Status",
        ]

        row = [cell.value for cell in sheet[2]]
        assert row[0] == 1
        assert row[1] == "Click Submit"
        assert row[3] == "84%"
        assert row[6] == "GENERATED"


def test_filename_derived_from_source_document():
    with tempfile.TemporaryDirectory() as tmp_dir:
        output_path = export_to_excel(_sample_result(), tmp_dir)
        assert os.path.basename(output_path) == "sample_test_steps.xlsx"


def test_creates_output_dir_if_missing():
    with tempfile.TemporaryDirectory() as tmp_dir:
        nested = os.path.join(tmp_dir, "reports", "run1")
        output_path = export_to_excel(_sample_result(), nested)
        assert os.path.exists(output_path)


def test_control_characters_from_garbled_ocr_do_not_crash_export():
    # Regression: openpyxl's IllegalCharacterError used to abort the
    # entire export over a single cell containing NUL/control bytes,
    # which garbled OCR from a low-quality screenshot can produce.
    result = GenerationResult(
        source_document="doc.docx",
        steps=[
            TestStep(
                step_no=1,
                action="Click Submit\x00 and verify\x0b\x1f the page",
                expected_result="fine",
                confidence=0.9,
            )
        ],
    )
    with tempfile.TemporaryDirectory() as tmp_dir:
        output_path = export_to_excel(result, tmp_dir)
        wb = load_workbook(output_path)
        action_cell = wb.active.cell(row=2, column=2).value
        assert "\x00" not in action_cell
        assert "Click Submit" in action_cell
